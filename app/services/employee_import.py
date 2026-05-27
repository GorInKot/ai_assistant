"""Импорт сотрудников из CSV/XLSX.

Поддерживаемые колонки (заголовки строкой 1, регистр и пробелы игнорируются):
  - email (обязательно)
  - full_name / фио (обязательно)
  - position / должность
  - division / филиал
  - subdivision / подразделение / пу
  - phone / телефон
  - responsibility_areas / области (через ';' или ',') — slug-и областей

Стратегия: upsert по email. Если email уже есть — обновляем поля
(пустые значения в файле НЕ перезаписывают существующие — это безопасно
для частичных листов).
"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass

from sqlalchemy.orm import Session

from app.db import Employee, Responsibility, ResponsibilityArea


EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


HEADER_ALIASES: dict[str, str] = {
    # email
    "email": "email",
    "e-mail": "email",
    "почта": "email",
    # full_name
    "full_name": "full_name",
    "fullname": "full_name",
    "fio": "full_name",
    "фио": "full_name",
    "имя": "full_name",
    # position
    "position": "position",
    "должность": "position",
    # division
    "division": "division",
    "филиал": "division",
    # subdivision
    "subdivision": "subdivision",
    "подразделение": "subdivision",
    "пу": "subdivision",
    "ауп": "subdivision",
    # phone
    "phone": "phone",
    "телефон": "phone",
    # areas
    "responsibility_areas": "areas",
    "области": "areas",
    "ответственность": "areas",
    "areas": "areas",
}


@dataclass
class ImportSummary:
    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: list[str] = None  # type: ignore[assignment]

    def __post_init__(self):
        if self.errors is None:
            self.errors = []


def _norm_header(value: str) -> str:
    return (value or "").strip().lower().replace(" ", "_")


def _parse_rows_from_csv(content: bytes) -> tuple[list[str], list[list[str]]]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return [], []
    return rows[0], rows[1:]


def _parse_rows_from_xlsx(content: bytes) -> tuple[list[str], list[list[str]]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl не установлен. pip install openpyxl") from exc

    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    sheet = workbook.active
    if sheet is None:
        return [], []

    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []

    headers = ["" if v is None else str(v) for v in header_row]
    data_rows: list[list[str]] = []
    for raw in rows_iter:
        data_rows.append(["" if v is None else str(v) for v in raw])
    return headers, data_rows


def import_employees(
    db: Session,
    *,
    content: bytes,
    filename: str,
) -> ImportSummary:
    summary = ImportSummary()

    lower_name = (filename or "").lower()
    if lower_name.endswith(".xlsx"):
        headers, data_rows = _parse_rows_from_xlsx(content)
    elif lower_name.endswith(".csv") or lower_name.endswith(".txt"):
        headers, data_rows = _parse_rows_from_csv(content)
    else:
        # Пробуем как CSV (наиболее частый случай прямого экспорта из Excel).
        headers, data_rows = _parse_rows_from_csv(content)

    if not headers:
        summary.errors.append("Файл пуст или нет заголовка")
        return summary

    # Маппим колонки на канонические имена.
    col_map: dict[int, str] = {}
    for idx, header in enumerate(headers):
        canonical = HEADER_ALIASES.get(_norm_header(header))
        if canonical:
            col_map[idx] = canonical

    if "email" not in col_map.values() or "full_name" not in col_map.values():
        summary.errors.append(
            "В заголовке должны быть колонки email и full_name (или фио). "
            f"Найдено: {[h for h in headers if h]}"
        )
        return summary

    # Подгружаем все известные slug-и областей для проверки.
    area_by_slug = {a.slug: a for a in db.query(ResponsibilityArea).all()}

    for row_idx, row in enumerate(data_rows, start=2):
        record: dict[str, str] = {}
        for col_idx, canonical in col_map.items():
            if col_idx < len(row):
                record[canonical] = (row[col_idx] or "").strip()

        email = record.get("email", "").lower()
        full_name = record.get("full_name", "")
        if not email and not full_name:
            summary.skipped += 1
            continue

        if not email or not EMAIL_RE.match(email):
            summary.errors.append(f"Строка {row_idx}: некорректный email '{email}'")
            summary.skipped += 1
            continue
        if not full_name:
            summary.errors.append(f"Строка {row_idx}: пустой full_name")
            summary.skipped += 1
            continue

        # Upsert по email.
        existing = db.query(Employee).filter(Employee.email == email).first()
        if existing:
            _apply_employee_updates(existing, record)
            updated = True
        else:
            existing = Employee(
                email=email,
                full_name=full_name,
                position=record.get("position") or None,
                division=record.get("division") or None,
                subdivision=record.get("subdivision") or None,
                phone=record.get("phone") or None,
                is_active=True,
            )
            db.add(existing)
            db.flush()
            updated = False

        # Обновляем области (только если поле явно указано в файле).
        if "areas" in record:
            area_slugs_raw = record["areas"]
            slugs = [s.strip() for s in re.split(r"[;,]", area_slugs_raw) if s.strip()]
            unknown = [slug for slug in slugs if slug not in area_by_slug]
            if unknown:
                summary.errors.append(
                    f"Строка {row_idx} ({email}): неизвестные области {unknown}"
                )
            valid_areas = [area_by_slug[s] for s in slugs if s in area_by_slug]
            _sync_global_areas(db, existing, valid_areas)

        if updated:
            summary.updated += 1
        else:
            summary.created += 1

    db.commit()
    return summary


def _apply_employee_updates(emp: Employee, record: dict[str, str]) -> None:
    # Перезаписываем только те поля, что явно есть в файле и непустые.
    if record.get("full_name"):
        emp.full_name = record["full_name"]
    if record.get("position"):
        emp.position = record["position"]
    if record.get("division"):
        emp.division = record["division"]
    if record.get("subdivision"):
        emp.subdivision = record["subdivision"]
    if record.get("phone"):
        emp.phone = record["phone"]
    # is_active при импорте не меняем — деактивация через UI/API.


def _sync_global_areas(db: Session, emp: Employee, areas: list[ResponsibilityArea]) -> None:
    """Синхронизирует глобальные responsibilities (scope=NULL) — для импорта.

    Импорт работает с глобальным уровнем; точечные scope (по подразделениям)
    настраиваются отдельно через UI.
    """
    desired_ids = {a.id for a in areas}
    current_global = [
        r for r in emp.responsibilities
        if r.scope_division is None and r.scope_subdivision is None
    ]
    current_ids = {r.area_id for r in current_global}

    for resp in current_global:
        if resp.area_id not in desired_ids:
            db.delete(resp)

    for area in areas:
        if area.id not in current_ids:
            db.add(Responsibility(
                employee_id=emp.id,
                area_id=area.id,
                scope_division=None,
                scope_subdivision=None,
                is_primary=True,
            ))
