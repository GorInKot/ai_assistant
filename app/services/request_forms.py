"""Автозаполнение шаблонов заявок на доступ к системам (Этап 7).

Пользователь свободным текстом в чате описывает данные сотрудника и просит
оформить заявку на конкретную систему. Ассистент определяет систему, через
LLM извлекает нужные поля, заполняет xlsx-шаблон и отдаёт готовый файл.

Шаблоны лежат в `app/form_templates/` и НЕ изменяются — заполняется копия в
памяти, файл на сервере не сохраняется (как и в Этапе 6).
"""

from __future__ import annotations

import base64
import io
import re
import zipfile
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Callable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side

from app.schemas import AskAttachment


TEMPLATES_DIR = Path(__file__).resolve().parent.parent / "form_templates"

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Глаголы-триггеры: пользователь именно просит ОФОРМИТЬ заявку, а не задаёт
# вопрос про систему. Без такого глагола интент не срабатывает — иначе обычный
# QA-вопрос «что за лаборатория ИИ?» ушёл бы в автозаполнение.
_FILL_VERBS = (
    "заполни",
    "заполнить",
    "оформи",
    "оформить",
    "сформируй",
    "сформировать",
    "сгенерируй",
    "сгенерировать",
    "создай заявк",
    "создать заявк",
    "состав",  # составь/составить
    "подготов",  # подготовь/подготовить
    "сделай заявк",
    "сделать заявк",
)

_MONTHS_GENITIVE = {
    1: "января",
    2: "февраля",
    3: "марта",
    4: "апреля",
    5: "мая",
    6: "июня",
    7: "июля",
    8: "августа",
    9: "сентября",
    10: "октября",
    11: "ноября",
    12: "декабря",
}


@dataclass
class FormField:
    key: str
    title: str
    required: bool = False
    description: str = ""


@dataclass
class FormSpec:
    key: str
    title: str
    template: str  # имя файла в form_templates/ (пусто, если строится с нуля)
    keywords: tuple[str, ...]  # триггеры названия системы
    fields: list[FormField]
    filler: Callable[["FormSpec", dict], bytes]
    output_name: Callable[[dict], str]
    # Коллективные заявки (напр. ВКД) перечисляют нескольких пользователей.
    # Для них `fields` описывают поля ОДНОГО пользователя, а filler получает
    # {"users": [ {поля}, ... ]}.
    collective: bool = False

    def fields_for_llm(self) -> list[dict]:
        return [
            {"key": f.key, "title": f.title, "description": f.description}
            for f in self.fields
        ]

    def missing_required(self, values: dict) -> list[FormField]:
        return [f for f in self.fields if f.required and not values.get(f.key)]


@dataclass
class FormResult:
    answer: str
    attachment: AskAttachment | None = None


# ----------------------------- утилиты заполнения ------------------------------


def _template_path(name: str) -> Path:
    path = TEMPLATES_DIR / name
    if not path.exists():
        raise RuntimeError(f"Шаблон заявки не найден на сервере: {name}")
    return path


def _join_person(parts: list[str | None], *, phone_prefix: str = "") -> str:
    """Склейка «ФИО, должность, телефон», пустые части пропускаются."""
    out: list[str] = []
    for idx, part in enumerate(parts):
        value = (part or "").strip()
        if not value:
            continue
        if idx == len(parts) - 1 and phone_prefix and out:
            out.append(f"{phone_prefix}{value}")
        else:
            out.append(value)
    return ", ".join(out)


_INLINE_CELL_RE = re.compile(
    r'<c\b([^>]*?)\bt="inlineStr"([^>]*)>\s*<is>(.*?)</is>\s*</c>',
    re.DOTALL,
)
_SHARED_CT = (
    '<Override PartName="/xl/sharedStrings.xml" '
    'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>'
)
_SHARED_REL_TYPE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings"
)


def _to_shared_strings(data: bytes) -> bytes:
    """Конвертирует inline-строки openpyxl в таблицу общих строк.

    openpyxl (3.1.x) пишет строки как inlineStr и не создаёт sharedStrings.xml.
    Excel/Google Sheets это читают, а Apple Numbers — нет. Переписываем книгу
    с `xl/sharedStrings.xml`, чтобы файл открывался везде. Поддерживаются
    простые строки `<is><t>…</t></is>` без rich-text (в наших шаблонах так и есть)
    — на всякий случай при наличии runs (`<is><r>`) конвертацию пропускаем.
    """
    with zipfile.ZipFile(io.BytesIO(data)) as zin:
        names = zin.namelist()
        members = {name: zin.read(name) for name in names}

    if "xl/sharedStrings.xml" in members:
        return data  # уже shared — ничего не делаем

    unique: list[str] = []
    index_by_inner: dict[str, int] = {}
    total = 0

    def _replace(match: re.Match) -> str:
        nonlocal total
        attrs = (match.group(1) + match.group(2)).rstrip()
        inner = match.group(3)
        idx = index_by_inner.get(inner)
        if idx is None:
            idx = len(unique)
            index_by_inner[inner] = idx
            unique.append(inner)
        total += 1
        return f'<c{attrs} t="s"><v>{idx}</v></c>'

    sheet_names = [n for n in names if n.startswith("xl/worksheets/") and n.endswith(".xml")]
    touched = False
    for name in sheet_names:
        xml = members[name].decode("utf-8")
        if "<is><r>" in xml:  # rich-text inline — не трогаем, отдаём как есть
            return data
        new_xml, count = _INLINE_CELL_RE.subn(_replace, xml)
        if count:
            members[name] = new_xml.encode("utf-8")
            touched = True

    if not touched:
        return data

    sst = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        f'count="{total}" uniqueCount="{len(unique)}">'
        + "".join(f"<si>{inner}</si>" for inner in unique)
        + "</sst>"
    )
    members["xl/sharedStrings.xml"] = sst.encode("utf-8")

    # Регистрируем часть в [Content_Types].xml.
    ct = members["[Content_Types].xml"].decode("utf-8")
    if "sharedStrings.xml" not in ct:
        ct = ct.replace("</Types>", _SHARED_CT + "</Types>")
        members["[Content_Types].xml"] = ct.encode("utf-8")

    # Добавляем relationship в xl/_rels/workbook.xml.rels.
    rels_name = "xl/_rels/workbook.xml.rels"
    rels = members[rels_name].decode("utf-8")
    existing_ids = [int(n) for n in re.findall(r'Id="rId(\d+)"', rels)]
    new_id = f"rId{(max(existing_ids) + 1) if existing_ids else 1}"
    rel = f'<Relationship Id="{new_id}" Type="{_SHARED_REL_TYPE}" Target="sharedStrings.xml"/>'
    rels = rels.replace("</Relationships>", rel + "</Relationships>")
    members[rels_name] = rels.encode("utf-8")

    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        for name in members:
            zout.writestr(name, members[name])
    return out.getvalue()


def _safe_filename(stem: str) -> str:
    cleaned = stem
    for ch in '/\\:*?"<>|':
        cleaned = cleaned.replace(ch, "-")
    cleaned = " ".join(cleaned.split())
    return cleaned or "Заявка"


# --------------------- заполнение «родного» Excel-пакета ----------------------
#
# openpyxl при load+save переписывает весь пакет в свой диалект (теряет
# printerSettings, worksheets/_rels, урезает namespaces). Excel и Google Sheets
# это читают, а Apple Numbers — нет. Поэтому для заявок патчим ИСХОДНЫЙ
# Excel-пакет напрямую: меняем только нужные ячейки и sharedStrings, всё
# остальное оставляем байт-в-байт. Так файл остаётся максимально «родным».


def _xml_escape(value: str) -> str:
    return value.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _set_cell(sheet_xml: str, ref: str, body: str, *, t: str | None) -> str:
    """Заменяет содержимое ячейки `ref`, сохраняя её стиль (атрибут s)."""
    pat = re.compile(rf'<c r="{re.escape(ref)}"([^>]*?)(/>|>.*?</c>)', re.DOTALL)
    m = pat.search(sheet_xml)
    if not m:
        raise RuntimeError(f"Ячейка {ref} не найдена в шаблоне заявки")
    style_match = re.search(r'\ss="(\d+)"', m.group(1))
    style = f' s="{style_match.group(1)}"' if style_match else ""
    t_attr = f' t="{t}"' if t else ""
    repl = f'<c r="{ref}"{style}{t_attr}>{body}</c>'
    return sheet_xml[: m.start()] + repl + sheet_xml[m.end() :]


def _fill_template_raw(
    template_name: str,
    string_cells: dict[str, str],
    number_cells: dict[str, int],
) -> bytes:
    """Заполняет xlsx-шаблон, не пересобирая пакет (для совместимости с Numbers)."""
    with zipfile.ZipFile(_template_path(template_name)) as zin:
        order = zin.namelist()
        members = {name: zin.read(name) for name in order}

    sst_name = "xl/sharedStrings.xml"
    sst = members[sst_name].decode("utf-8")
    si_count = len(re.findall(r"<si[ >]", sst))

    # Добавляем строковые значения в конец таблицы общих строк.
    new_si: list[str] = []
    ref_to_idx: dict[str, int] = {}
    for ref, value in string_cells.items():
        ref_to_idx[ref] = si_count + len(new_si)
        new_si.append(f'<si><t xml:space="preserve">{_xml_escape(value)}</t></si>')

    if new_si:
        total = si_count + len(new_si)
        sst = sst.replace("</sst>", "".join(new_si) + "</sst>")
        sst = re.sub(r'\bcount="\d+"', f'count="{total}"', sst, count=1)
        sst = re.sub(r'uniqueCount="\d+"', f'uniqueCount="{total}"', sst, count=1)
        members[sst_name] = sst.encode("utf-8")

    sheet_name = "xl/worksheets/sheet1.xml"
    sheet = members[sheet_name].decode("utf-8")
    for ref, idx in ref_to_idx.items():
        sheet = _set_cell(sheet, ref, f"<v>{idx}</v>", t="s")
    for ref, num in number_cells.items():
        sheet = _set_cell(sheet, ref, f"<v>{num}</v>", t=None)
    members[sheet_name] = sheet.encode("utf-8")

    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        for name in order:  # сохраняем исходный порядок частей пакета
            zout.writestr(name, members[name])
    return out.getvalue()


# ------------------------------- Лаборатория ИИ -------------------------------


def _fill_lab_ai(spec: FormSpec, v: dict) -> bytes:
    """Заполняет Шаблон_Лаборатория.xlsx по логике VBA-макроса LaboratoryAI_3_1.

    Патчим исходный Excel-пакет напрямую (не через openpyxl save), чтобы файл
    открывался и в Apple Numbers.
    """
    division = (v.get("division") or "").strip()
    today = date.today()
    month = _MONTHS_GENITIVE[today.month]

    string_cells = {
        # Подразделение — в две ячейки (как в макросе: F15 и H47).
        "F15": division,
        "H47": division,
        # F16: ФИО, должность, телефон работника.
        "F16": _join_person(
            [v.get("employee_name"), v.get("employee_position"), v.get("employee_phone")]
        ),
        # F17: ФИО, должность, тел. руководителя.
        "F17": _join_person(
            [v.get("manager_name"), v.get("manager_position"), v.get("manager_phone")],
            phone_prefix="тел.",
        ),
        # F18: имя ключа ПКЗИ.
        "F18": (v.get("pkzi") or "").strip(),
        # Месяц прописью (период «с … по …»).
        "H31": month,
        "M31": month,
        # ФИО внизу: работник (B43) и руководитель (B50).
        "B43": (v.get("employee_name") or "").strip(),
        "B50": (v.get("manager_name") or "").strip(),
    }
    number_cells = {
        # Дата: текущая, период «с» текущего года «по» год+1 (как в макросе).
        "F31": today.day,
        "I31": today.year,
        "K31": today.day,
        "N31": today.year + 1,
    }
    return _fill_template_raw(spec.template, string_cells, number_cells)


def _lab_ai_output_name(v: dict) -> str:
    fio = (v.get("employee_name") or "").strip()
    return f"Заявка_{_safe_filename(fio)}.xlsx" if fio else "Заявка_Лаборатория.xlsx"


LAB_AI = FormSpec(
    key="lab_ai",
    title="Лаборатория ИИ",
    template="lab_ai.xlsx",
    keywords=(
        "лаборатори",
        "лаб ии",
        "лаб. ии",
        "стенд ии",
        "стенда ии",
        "стенд искусств",
        "искусственн",
    ),
    fields=[
        FormField("division", "Наименование структурного подразделения", required=True),
        FormField("employee_name", "ФИО работника", required=True),
        FormField("employee_position", "Должность работника"),
        FormField("employee_phone", "Телефон работника"),
        FormField("manager_name", "ФИО непосредственного руководителя", required=True),
        FormField("manager_position", "Должность руководителя"),
        FormField("manager_phone", "Телефон руководителя"),
        FormField("pkzi", "Имя ключа ПКЗИ", required=True),
    ],
    filler=_fill_lab_ai,
    output_name=_lab_ai_output_name,
)


# ------------------------------------- ВКД ------------------------------------
#
# Групповая заявка на ИС «Виртуальные комнаты данных». В отличие от Лаборатории
# ИИ строится С НУЛЯ (готового шаблона нет — воспроизводим VBA-макрос
# `VKD_ГрупповаяЗаявка`) и собирает в одну заявку НЕСКОЛЬКИХ пользователей.
#
# Организационные константы (наименование ВКД, компания, подписанты, основание)
# зашиты как в макросе — по аналогии с подписантами в Лаборатории ИИ. Меняются
# здесь же при необходимости.

VKD_TITLE = "ЗАЯВКА НА УПРАВЛЕНИЕ ДОСТУПОМ К СИСТЕМЕ ПОЛЬЗОВАТЕЛЕЙ ВКД"
VKD_SUBTITLE = "ИС «Виртуальные комнаты данных»"
VKD_OWNER = "Кулешов А.П."  # Владелец ВКД (Делегат ВКД)
VKD_IB_DEPARTMENT = (
    "Структурное подразделение информационной безопасности ПАО «НК «Роснефть»"
)
VKD_NAME = "РНСК"  # Наименование ВКД (столбец B)
VKD_COMPANY = 'ООО «РН-СтройКонтроль»'  # Компания пользователя
VKD_ACCESS_REASON = "Исполнение служебных обязанностей, обмен информацией с филиалами"
VKD_ACCOUNT_PREFIX = "ROSNEFT\\"  # домен учётной записи
VKD_APPROVER_HEAD = "Кулешов А.П."  # руководитель СП
VKD_APPROVER_IB = "Литвинов В.А."  # ИБ ОГ
VKD_CHECKBOX_ON = "☑"  # ☑
VKD_CHECKBOX_OFF = "☐"  # ☐

_TNR = "Times New Roman"
_THIN = Side(style="thin")
_BOX = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_UNDERLINE = Border(bottom=_THIN)


def _f(size: int = 12, *, bold: bool = False, italic: bool = False) -> Font:
    return Font(name=_TNR, size=size, bold=bold, italic=italic)


def _fill_vkd(spec: FormSpec, v: dict) -> bytes:
    """Строит групповую заявку ВКД с нуля по логике макроса VKD_ГрупповаяЗаявка."""
    users: list[dict] = v.get("users") or []

    wb = Workbook()
    ws = wb.active
    ws.title = "Заявка_ВКД"

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def put(ref: str, value, *, font: Font | None = None, align: Alignment | None = None,
            border: Border | None = None, merge: str | None = None) -> None:
        if merge:
            ws.merge_cells(merge)
        cell = ws[ref]
        cell.value = value
        cell.font = font or _f()
        cell.alignment = align or center
        if border is not None:
            # При merge границу ставим на все ячейки диапазона, иначе только на одну.
            if merge:
                for row in ws[merge]:
                    for c in row:
                        c.border = border
            else:
                cell.border = border

    # ----- шапка документа -----
    put("A1", VKD_TITLE, font=_f(14, bold=True), merge="A1:H1")
    put("A2", VKD_SUBTITLE, font=_f(14), merge="A2:H2")
    put("A4", "СОГЛАСОВАНО", font=_f(11, bold=True, italic=True), align=left,
        merge="A4:B4")

    put("A5", "Владелец ВКД (Делегат ВКД)", align=left, merge="A5:D5")
    put("E5", VKD_OWNER, border=_UNDERLINE)
    put("F5", "/                                 /", border=_UNDERLINE)
    put("G5", "«___» _______ 20__г. ", merge="G5:H5")

    put("A6", VKD_IB_DEPARTMENT, align=left, merge="A6:D7")
    put("E7", None, border=_UNDERLINE)
    put("F7", "/                                 /", border=_UNDERLINE)
    put("G7", "«___» _______ 20__г. ", merge="G7:H7")

    # ----- действия с учётной записью (чекбоксы) -----
    put("A9", "Действия с учетной записью:\n(выберите один пункт)",
        border=_BOX, merge="A9:B11")
    put("C9", f"{VKD_CHECKBOX_ON} Предоставление доступа\nПользователям ВКД",
        border=_BOX, merge="C9:C11")
    put("D9", f"{VKD_CHECKBOX_OFF} Изменение\nатрибутов УЗ\nПользователей ВКД",
        border=_BOX, merge="D9:D11")
    put("E9", f"{VKD_CHECKBOX_OFF} Прекращение (изъятие) доступа\nПользователей ВКД",
        border=_BOX, merge="E9:E11")
    put("F9", f"{VKD_CHECKBOX_OFF} Блокировка УЗ в ИС ВКД\nПользователей ВКД",
        border=_BOX, merge="F9:G11")
    put("H9", f"{VKD_CHECKBOX_OFF} Согласование полномочий, присвоенных ранее в "
        "экстренном порядке", border=_BOX, merge="H9:H11")

    # ----- шапка таблицы -----
    put("A12", " ИНФОРМАЦИЯ О ПОЛЬЗОВАТЕЛЯХ ВКД", font=_f(12, bold=True),
        border=_BOX, merge="A12:H12")

    put("A13", "№", border=_BOX, merge="A13:A14")
    put("B13", "Наименование ВКД", border=_BOX)
    put("B14", "(допускается указывать несколько ВКД, если у них один Владелец)",
        font=_f(8), border=_BOX)
    put("C13", "Пользователь ВКД", border=_BOX, merge="C13:D13")
    put("C14", "(поля заполняются полностью, без сокращений)", font=_f(8),
        border=_BOX, merge="C14:D14")
    put("E13", "Доступ предоставляется", border=_BOX)
    put("E14", "(выбрать один пункт для каждого сотрудника)", font=_f(8), border=_BOX)
    put("F13", "Имя учетной записи", border=_BOX)
    put("F14", "(в домене rosneft или vdr-domain.local с указанием домена)",
        font=_f(8), border=_BOX)
    put("G13", "Имя ключа", border=_BOX)
    put("G14", "ПКЗИ-КТ", font=_f(10), border=_BOX)
    put("H13", "Основание предоставления доступа партнерам/контрагентам", border=_BOX)
    put("H14", "(№ и дата документа Соглашения о конфиденциальности/ договора)",
        font=_f(10), border=_BOX)

    # ----- строки пользователей (по 6 строк на каждого) -----
    start_row = 15
    for idx, user in enumerate(users):
        r = start_row + idx * 6
        name = (user.get("employee_name") or "").strip()
        email = (user.get("email") or "").strip()
        account = (user.get("account") or "").strip()
        pkzi = (user.get("pkzi") or "").strip()

        put(f"A{r}", idx + 1, merge=f"A{r}:A{r + 5}")
        put(f"B{r}", VKD_NAME, merge=f"B{r}:B{r + 5}")

        # Колонка C — подписи, колонка D — значения.
        put(f"C{r}", "ФИО:", align=left)
        put(f"D{r}", name, font=_f(italic=True), align=left)
        put(f"C{r + 1}", "E-mail:", align=left)
        put(f"D{r + 1}", email, font=_f(italic=True), align=left)
        put(f"C{r + 2}", "Компания:", align=left)
        put(f"D{r + 2}", VKD_COMPANY, align=left)
        put(f"C{r + 3}", "Для партнера/контрагента:", font=_f(italic=True),
            merge=f"C{r + 3}:D{r + 3}")
        put(f"C{r + 4}", "Страна:", align=left)
        put(f"C{r + 5}", "Номер моб. телефона (для отправки SMS-кода)", align=left)

        put(f"E{r}", f"{VKD_CHECKBOX_ON}До срока действия ВКД",
            font=_f(bold=True), align=left, merge=f"E{r}:E{r + 2}")
        put(f"E{r + 3}", f"{VKD_CHECKBOX_OFF}До __. __. ____",
            align=Alignment(horizontal="left", vertical="top", wrap_text=True),
            merge=f"E{r + 3}:E{r + 5}")

        put(f"F{r}", f"{VKD_ACCOUNT_PREFIX}{account}" if account else "",
            merge=f"F{r}:F{r + 5}")
        put(f"G{r}", pkzi, merge=f"G{r}:G{r + 5}")
        put(f"H{r}", VKD_ACCESS_REASON, merge=f"H{r}:H{r + 5}")

        # Рамка вокруг всего блока пользователя.
        for row in ws[f"A{r}:H{r + 5}"]:
            for c in row:
                if c.border == Border():
                    c.border = _BOX

    # ----- блоки согласования под таблицей -----
    last_row = start_row + max(len(users), 1) * 6
    ce = last_row + 2

    put(f"A{ce}", "Согласовано руководителем СП (для ОГ – руководителем СП ОГ):",
        font=_f(bold=True, italic=True), align=left, merge=f"A{ce}:D{ce}")
    put(f"A{ce + 1}", "Заместитель генерального директора по развитию ",
        font=_f(bold=True, italic=True), align=left, border=_UNDERLINE,
        merge=f"A{ce + 1}:C{ce + 1}")
    put(f"B{ce + 2}", "Должность руководителя СП (СП ОГ)", font=_f(8),
        merge=f"B{ce + 2}:C{ce + 2}")
    put(f"B{ce + 3}", "(полностью, без сокращений)", font=_f(8),
        merge=f"B{ce + 3}:C{ce + 3}")
    put(f"E{ce + 1}", VKD_APPROVER_HEAD, font=_f(bold=True, italic=True),
        border=_UNDERLINE, merge=f"E{ce + 1}:F{ce + 1}")
    put(f"E{ce + 2}", "Фамилия Имя Отчество", font=_f(8), merge=f"E{ce + 2}:F{ce + 2}")
    put(f"E{ce + 3}", "(полностью, без сокращений)", font=_f(8),
        merge=f"E{ce + 3}:F{ce + 3}")
    put(f"H{ce + 1}", None, border=_UNDERLINE)
    put(f"H{ce + 2}", "Подпись", font=_f(8))

    cb = ce + 5
    put(f"A{cb}", "Согласовано ИБ ОГ (для ОГ):", font=_f(bold=True, italic=True),
        align=left, merge=f"A{cb}:D{cb}")
    put(f"A{cb + 1}", "Главный специалист, группа ИБ", font=_f(bold=True, italic=True),
        align=left, border=_UNDERLINE, merge=f"A{cb + 1}:C{cb + 1}")
    put(f"B{cb + 2}", "должность работника Информационной безопасности ОГ",
        font=_f(8), merge=f"B{cb + 2}:C{cb + 2}")
    put(f"B{cb + 3}", "(полностью, без сокращений)", font=_f(8),
        merge=f"B{cb + 3}:C{cb + 3}")
    put(f"E{cb + 1}", VKD_APPROVER_IB, font=_f(bold=True, italic=True),
        border=_UNDERLINE, merge=f"E{cb + 1}:F{cb + 1}")
    put(f"E{cb + 2}", "Фамилия Имя Отчество", font=_f(8), merge=f"E{cb + 2}:F{cb + 2}")
    put(f"E{cb + 3}", "(полностью, без сокращений)", font=_f(8),
        merge=f"E{cb + 3}:F{cb + 3}")
    put(f"H{cb + 1}", None, border=_UNDERLINE)
    put(f"H{cb + 2}", "Подпись", font=_f(8))

    # ----- ширины колонок (как в макросе) -----
    for col, width in zip("ABCDEFGH", (5, 25, 25, 30, 20, 20, 15, 30)):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _vkd_output_name(v: dict) -> str:
    return f"Заявка_ВКД_{date.today():%d.%m.%Y}.xlsx"


VKD = FormSpec(
    key="vkd",
    title="ВКД",
    template="",  # строится с нуля
    keywords=("вкд", "виртуальн", "комнат данных"),
    fields=[
        FormField("employee_name", "ФИО пользователя", required=True),
        FormField("email", "E-mail пользователя"),
        FormField("account", "Имя учётной записи (логин в домене)", required=True),
        FormField("pkzi", "Имя ключа ПКЗИ", required=True),
    ],
    filler=_fill_vkd,
    output_name=_vkd_output_name,
    collective=True,
)


# ------------------------------- реестр систем --------------------------------

FORMS: list[FormSpec] = [LAB_AI, VKD]


def detect_form_intent(message: str) -> FormSpec | None:
    """Определяет, просит ли пользователь оформить заявку на одну из систем.

    Требуется И глагол-триггер (оформи/заполни/…), И упоминание системы.
    Это намеренно консервативно, чтобы не перехватывать обычные вопросы.
    """
    text = message.lower()
    if not any(verb in text for verb in _FILL_VERBS):
        return None
    for spec in FORMS:
        if any(kw in text for kw in spec.keywords):
            return spec
    return None


def _missing_fields_answer(spec: FormSpec, missing: list[FormField], values: dict) -> str:
    have = [f for f in spec.fields if values.get(f.key)]
    lines = [f"Чтобы оформить заявку «{spec.title}», не хватает данных:"]
    lines += [f"- {f.title}" for f in missing]
    if have:
        lines.append("")
        lines.append("Уже распознано:")
        lines += [f"- {f.title}: {values[f.key]}" for f in have]
    lines.append("")
    lines.append("Допишите недостающее одним сообщением — и я сформирую файл.")
    return "\n".join(lines)


def _build_attachment(spec: FormSpec, content: bytes, values: dict) -> AskAttachment:
    return AskAttachment(
        filename=spec.output_name(values),
        content_base64=base64.b64encode(_to_shared_strings(content)).decode("ascii"),
        mime=XLSX_MIME,
    )


def _collective_missing_answer(spec: FormSpec, problems: list[tuple[int, dict, list[FormField]]]) -> str:
    lines = [f"Чтобы оформить групповую заявку «{spec.title}», не хватает данных:"]
    for num, person, missing in problems:
        who = person.get("employee_name") or f"пользователь №{num}"
        lines.append(f"- {who}: " + ", ".join(f.title for f in missing))
    lines.append("")
    lines.append("Допишите недостающее — и я сформирую файл.")
    return "\n".join(lines)


def _handle_collective(spec: FormSpec, message: str, llm) -> FormResult:
    people = llm.extract_people(message, spec.fields_for_llm())
    if not people:
        return FormResult(
            answer=(
                f"Для заявки «{spec.title}» не удалось распознать ни одного "
                "пользователя. Перечислите сотрудников и их данные "
                f"({', '.join(f.title for f in spec.fields if f.required)})."
            )
        )

    problems = [
        (idx, person, missing)
        for idx, person in enumerate(people, start=1)
        if (missing := spec.missing_required(person))
    ]
    if problems:
        return FormResult(answer=_collective_missing_answer(spec, problems))

    values = {"users": people}
    attachment = _build_attachment(spec, spec.filler(spec, values), values)

    lines = [
        f"✅ Групповая заявка «{spec.title}» сформирована. "
        f"Пользователей: {len(people)}.",
        "",
    ]
    for num, person in enumerate(people, start=1):
        parts = [person[f.key] for f in spec.fields if person.get(f.key)]
        lines.append(f"{num}. " + " · ".join(parts))
    lines.append("")
    lines.append(f"Файл «{attachment.filename}» приложен ниже — скачайте и проверьте.")
    return FormResult(answer="\n".join(lines), attachment=attachment)


def handle_form_request(spec: FormSpec, message: str, llm) -> FormResult:
    """Извлекает поля из текста и заполняет шаблон. Ошибки LLM → RuntimeError."""
    if spec.collective:
        return _handle_collective(spec, message, llm)

    values = llm.extract_fields(message, spec.fields_for_llm())

    missing = spec.missing_required(values)
    if missing:
        return FormResult(answer=_missing_fields_answer(spec, missing, values))

    content = _to_shared_strings(spec.filler(spec, values))
    filename = spec.output_name(values)
    attachment = AskAttachment(
        filename=filename,
        content_base64=base64.b64encode(content).decode("ascii"),
        mime=XLSX_MIME,
    )

    summary_lines = [f"✅ Заявка «{spec.title}» сформирована. Проверьте данные:"]
    for f in spec.fields:
        if values.get(f.key):
            summary_lines.append(f"- {f.title}: {values[f.key]}")
    summary_lines.append("")
    summary_lines.append(f"Файл «{filename}» приложен ниже — скачайте и при необходимости поправьте.")
    return FormResult(answer="\n".join(summary_lines), attachment=attachment)
