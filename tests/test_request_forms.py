"""Тесты автозаполнения шаблонов заявок (Этап 7)."""

from __future__ import annotations

import base64
import io
import zipfile
from datetime import date

import pytest
from openpyxl import load_workbook

import app.state as state
from app.services import request_forms as rf
from tests.conftest import ADMIN_EMAIL, ADMIN_PASSWORD, auth_headers


@pytest.fixture
def headers(client):
    return auth_headers(client, ADMIN_EMAIL, ADMIN_PASSWORD)


_FULL = {
    "division": "Отдел строительного контроля",
    "employee_name": "Иванов Иван Иванович",
    "employee_position": "Инженер",
    "employee_phone": "+7 999 123-45-67",
    "manager_name": "Петров Пётр Петрович",
    "manager_position": "Начальник отдела",
    "manager_phone": "+7 999 765-43-21",
    "pkzi": "ПКЗИ-КТ-12345",
}


class _StubLLM:
    """LLM, возвращающий заранее заданный набор полей."""

    def __init__(self, values: dict):
        self._values = values

    def extract_fields(self, text, fields):
        # Возвращаем только ключи, объявленные в спецификации.
        keys = {f["key"] for f in fields}
        return {k: self._values.get(k) for k in keys}


class _StubPeopleLLM:
    """LLM для коллективных заявок — возвращает заранее заданный список людей."""

    def __init__(self, people: list[dict]):
        self._people = people

    def extract_people(self, text, fields):
        keys = {f["key"] for f in fields}
        return [{k: p.get(k) for k in keys} for p in self._people]


_VKD_USERS = [
    {
        "employee_name": "Иванов Иван Иванович",
        "email": "ivanov@rn.ru",
        "account": "ivanov_ii",
        "pkzi": "ПКЗИ-КТ-111",
    },
    {
        "employee_name": "Петров Пётр Петрович",
        "email": "petrov@rn.ru",
        "account": "petrov_pp",
        "pkzi": "ПКЗИ-КТ-222",
    },
]


# ----------------------------- детект интента --------------------------------


def test_detect_requires_verb_and_system():
    assert rf.detect_form_intent("оформи заявку на доступ к лаборатории ИИ").key == "lab_ai"
    assert rf.detect_form_intent("заполни заявку на стенд ИИ для Иванова").key == "lab_ai"


def test_detect_ignores_questions():
    assert rf.detect_form_intent("что такое лаборатория ИИ?") is None
    assert rf.detect_form_intent("кто отвечает за стенд ИИ?") is None


def test_detect_ignores_unknown_system():
    # Глагол есть, но система пока не зарегистрирована.
    assert rf.detect_form_intent("заполни заявку на ЦУС") is None


def test_detect_vkd():
    assert rf.detect_form_intent("оформи заявку на ВКД для Иванова и Петрова").key == "vkd"
    assert rf.detect_form_intent("сформируй заявку на виртуальные комнаты данных").key == "vkd"
    assert rf.detect_form_intent("что такое ВКД?") is None


# ----------------------------- заполнение xlsx -------------------------------


def test_fill_lab_ai_writes_expected_cells():
    data = rf._fill_lab_ai(rf.LAB_AI, _FULL)
    ws = load_workbook(io.BytesIO(data)).active

    assert ws["F15"].value == _FULL["division"]
    assert ws["H47"].value == _FULL["division"]
    assert ws["F16"].value == "Иванов Иван Иванович, Инженер, +7 999 123-45-67"
    assert ws["F17"].value == "Петров Пётр Петрович, Начальник отдела, тел.+7 999 765-43-21"
    assert ws["F18"].value == "ПКЗИ-КТ-12345"
    assert ws["B43"].value == _FULL["employee_name"]
    assert ws["B50"].value == _FULL["manager_name"]

    today = date.today()
    assert ws["F31"].value == today.day
    assert ws["I31"].value == today.year
    assert ws["N31"].value == today.year + 1
    assert ws["H31"].value == rf._MONTHS_GENITIVE[today.month]


def test_fill_skips_empty_optional_parts():
    values = {**_FULL, "employee_position": None, "employee_phone": None}
    ws = load_workbook(io.BytesIO(rf._fill_lab_ai(rf.LAB_AI, values))).active
    # Без должности и телефона — только ФИО, без висящих запятых.
    assert ws["F16"].value == "Иванов Иван Иванович"


# ----------------------------- handle_form_request ---------------------------


def test_handle_full_data_produces_attachment():
    res = rf.handle_form_request(rf.LAB_AI, "оформи заявку на лабораторию ИИ", _StubLLM(_FULL))
    assert res.attachment is not None
    assert res.attachment.filename == "Заявка_Иванов Иван Иванович.xlsx"
    wb = load_workbook(io.BytesIO(base64.b64decode(res.attachment.content_base64)))
    assert wb.active["F18"].value == "ПКЗИ-КТ-12345"


def test_generated_file_uses_shared_strings_for_numbers():
    # Apple Numbers не читает inline-строки openpyxl — итоговый файл должен
    # содержать sharedStrings.xml и не содержать inlineStr.
    res = rf.handle_form_request(rf.LAB_AI, "оформи заявку на лабораторию ИИ", _StubLLM(_FULL))
    data = base64.b64decode(res.attachment.content_base64)
    z = zipfile.ZipFile(io.BytesIO(data))
    assert "xl/sharedStrings.xml" in z.namelist()
    assert b"inlineStr" not in z.read("xl/worksheets/sheet1.xml")
    # значения по-прежнему читаются
    assert load_workbook(io.BytesIO(data)).active["F18"].value == "ПКЗИ-КТ-12345"


def test_handle_missing_required_asks_for_data():
    partial = {"employee_name": "Иванов И.И.", "division": "Отдел"}  # нет руководителя и ПКЗИ
    res = rf.handle_form_request(rf.LAB_AI, "оформи заявку на лабораторию ИИ", _StubLLM(partial))
    assert res.attachment is None
    assert "не хватает" in res.answer.lower()
    assert "ключ" in res.answer.lower()  # ПКЗИ среди недостающих


# ------------------------------------- ВКД -----------------------------------


def test_fill_vkd_writes_users_and_constants():
    data = rf._fill_vkd(rf.VKD, {"users": _VKD_USERS})
    ws = load_workbook(io.BytesIO(data)).active

    assert ws["A1"].value == rf.VKD_TITLE
    # Первый пользователь (строки 15-20).
    assert ws["A15"].value == 1
    assert ws["B15"].value == rf.VKD_NAME
    assert ws["D15"].value == "Иванов Иван Иванович"
    assert ws["D16"].value == "ivanov@rn.ru"
    assert ws["D17"].value == rf.VKD_COMPANY
    assert ws["F15"].value == "ROSNEFT\\ivanov_ii"
    assert ws["G15"].value == "ПКЗИ-КТ-111"
    assert ws["H15"].value == rf.VKD_ACCESS_REASON
    # Второй пользователь — следующие 6 строк.
    assert ws["A21"].value == 2
    assert ws["D21"].value == "Петров Пётр Петрович"
    # Подписанты согласования.
    assert ws["E30"].value == rf.VKD_APPROVER_HEAD
    assert ws["E35"].value == rf.VKD_APPROVER_IB


def test_vkd_generated_file_numbers_compatible():
    res = rf.handle_form_request(rf.VKD, "оформи заявку на ВКД", _StubPeopleLLM(_VKD_USERS))
    data = base64.b64decode(res.attachment.content_base64)
    z = zipfile.ZipFile(io.BytesIO(data))
    assert "xl/sharedStrings.xml" in z.namelist()
    assert b"inlineStr" not in z.read("xl/worksheets/sheet1.xml")


def test_vkd_handle_produces_attachment_and_summary():
    res = rf.handle_form_request(rf.VKD, "оформи заявку на ВКД", _StubPeopleLLM(_VKD_USERS))
    assert res.attachment is not None
    assert res.attachment.filename.startswith("Заявка_ВКД_")
    assert "Пользователей: 2" in res.answer
    assert "Иванов Иван Иванович" in res.answer


def test_vkd_missing_required_lists_user():
    incomplete = [
        _VKD_USERS[0],
        {"employee_name": "Сидоров С.С.", "email": None, "account": None, "pkzi": None},
    ]
    res = rf.handle_form_request(rf.VKD, "оформи заявку на ВКД", _StubPeopleLLM(incomplete))
    assert res.attachment is None
    assert "Сидоров С.С." in res.answer
    assert "не хватает" in res.answer.lower()


def test_vkd_no_users_recognized():
    res = rf.handle_form_request(rf.VKD, "оформи заявку на ВКД", _StubPeopleLLM([]))
    assert res.attachment is None
    assert "ни одного" in res.answer.lower()


# ----------------------------- интеграция /api/ask ---------------------------


def test_ask_endpoint_fills_form(client, headers, monkeypatch):
    monkeypatch.setattr(state.llm_service, "extract_fields", lambda text, fields: dict(_FULL))
    resp = client.post(
        "/api/ask",
        json={"question": "оформи заявку на доступ к лаборатории ИИ для Иванова"},
        headers=headers,
    )
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert data["attachment"] is not None
    assert data["attachment"]["filename"].startswith("Заявка_")


def test_ask_endpoint_missing_data_no_attachment(client, headers, monkeypatch):
    monkeypatch.setattr(
        state.llm_service,
        "extract_fields",
        lambda text, fields: {"employee_name": "Иванов И.И."},
    )
    resp = client.post(
        "/api/ask",
        json={"question": "оформи заявку на лабораторию ИИ"},
        headers=headers,
    )
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert data["attachment"] is None
    assert "не хватает" in data["answer"].lower()
